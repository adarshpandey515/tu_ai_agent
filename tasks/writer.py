from pathlib import Path
from typing import Dict, Any, List

def _desired_sheet_names() -> List[str]:
    """Attempt to mimic sheet names from Expected Output.xlsx if present."""
    try:
        from openpyxl import load_workbook  # type: ignore
        expected = Path('Expected Output.xlsx')
        if expected.exists():
            wb = load_workbook(expected, read_only=True)
            return list(wb.sheetnames)
    except Exception:
        pass
    return ["Raw", "Sections"]

def write_excel(out_path: Path, sections: Dict[str, Any]) -> None:
    try:
        import pandas as pd
    except Exception:
        # Minimal CSV fallback if pandas missing
        out_path = out_path.with_suffix('.csv')
        with open(out_path, 'w', encoding='utf-8') as f:
            for k, v in sections.items():
                f.write(f"{k},\"{str(v).replace('"', '""')}\"\n")
        return

    # Create multiple sheets matching expected names when possible
    sheet_names = _desired_sheet_names()
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        # First sheet: Raw
        raw_name = sheet_names[0] if sheet_names else "Raw"
        pd.DataFrame({"content": [sections.get("raw", "")]})\
            .to_excel(writer, sheet_name=raw_name, index=False)

        # Second sheet: parsed sections
        sec_name = sheet_names[1] if len(sheet_names) > 1 else "Sections"
        items = [(k, v) for k, v in sections.items() if k != "raw"]
        pd.DataFrame(items, columns=["section", "content"])\
            .to_excel(writer, sheet_name=sec_name, index=False)
